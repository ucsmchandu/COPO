from django.shortcuts import render,redirect,get_object_or_404
from django.contrib.auth import authenticate, login, logout
from .forms import RegisterForm
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from .models import Course, CO, PO, COPOMapping, COAttainment, Student, StudentTotal, StudentMark, COAggregateScore, POAttainment
from django.contrib.auth.models import User
from openpyxl import load_workbook
from collections import defaultdict
from django.http import HttpResponse, Http404
from statistics import mode, StatisticsError

# Create your views here.
def home(request):
    return render(request, 'home.html')


def register(request):
    if request.method == 'POST':
        form = RegisterForm(request.POST)
        if form.is_valid():
            user = form.save()
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password1')
            user = authenticate(username=username, password=password)
            if user is not None:
                login(request, user)
                return render(request, 'home.html', {'message': 'Registration successful!'})
    else:
        form = RegisterForm()
    return render(request, 'register.html', {'form': form})



def login_view(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return render(request, 'home.html', {
                'message': 'Login successful!',
                'user': user
            })
        else:
            return render(request, 'login.html', {
                'error': 'Invalid credentials'
            })
    return render(request, 'login.html')


def logout_view(request):
    logout(request)
    return redirect('login') 

def add_course(request):
    if request.method == 'POST':
        course_code = request.POST.get('course_code')
        course_name = request.POST.get('course_name')
        user_id = request.POST.get('user')  # 👈 get from form

        try:
            user = User.objects.get(id=user_id)
            Course.objects.create(
                code=course_code,
                name=course_name,
                user=user  # ✅ set user from form
            )
            return redirect('add_course')  # or show success page
        except User.DoesNotExist:
            return render(request, 'add_course.html', {
                'error': 'Invalid user selected.'
            })

    users = User.objects.all()
    return render(request, 'add_course.html', {
        'users': users,
        'user': request.user
    })


def courses(request):
    if request.user.is_authenticated:
        courses = Course.objects.filter(user=request.user)
        return render(request, 'courses.html', {'courses': courses})
    else:
        return redirect('login')  # Redirect to login if not authenticated
    
@login_required
def add_co(request):
    if request.method == 'POST':
        course_id = request.POST.get('course')
        co_number = request.POST.get('co_number')
        co_description = request.POST.get('description')
        max_marks = request.POST.get('max_marks')  # <-- get from form

        try:
            course = Course.objects.get(id=course_id, user=request.user)

            if CO.objects.filter(course=course, number=co_number).exists():
                messages.warning(request, "This CO already exists for the selected course.")
            else:
                CO.objects.create(
                    course=course,
                    number=co_number,
                    description=co_description,
                    max_score=max_marks  # <-- save to db
                )
                messages.success(request, "Course Outcome added successfully!")
            return redirect('add_co')

        except Course.DoesNotExist:
            messages.error(request, "Invalid course selected.")

    courses = Course.objects.filter(user=request.user)
    return render(request, 'add_co.html', {'courses': courses})

def add_po(request):
    if request.method == 'POST':
        po_number = request.POST.get('po_number')
        po_description = request.POST.get('po_description')

        PO.objects.create(
            number=po_number,
            description=po_description
        )
        return redirect('add_po')  # or show success page

    return render(request, 'add_po.html')


@login_required
def add_mapping(request):
    if request.method == 'POST':
        processed = False
        course_id = request.POST.get('course')
        if not course_id:
            messages.error(request, "Course selection is required for mapping.")
            return redirect('add_mapping')

        # Excel upload
        excel_file = request.FILES.get('excel_file')
        if excel_file:
            try:
                wb = load_workbook(excel_file)
                sheet = wb.active

                # Read PO headers from first row (skip first column)
                po_headers = [str(cell.value).strip().upper() for cell in sheet[1][1:]]

                # Prepare PO and CO objects for fast lookup
                po_map = {po.number.upper(): po for po in PO.objects.all()}
                co_map = {co.number.upper(): co for co in CO.objects.filter(course_id=course_id)}

                for row in sheet.iter_rows(min_row=2, values_only=True):
                    co_key = str(row[0]).strip().upper()
                    co = co_map.get(co_key)

                    if not co:
                        continue

                    for i, level in enumerate(row[1:]):
                        po_key = po_headers[i]
                        po = po_map.get(po_key)

                        if po and level:
                            COPOMapping.objects.update_or_create(
                                co=co,
                                po=po,
                                defaults={'level': int(level)}
                            )
                messages.success(request, "CO-PO mappings updated from Excel file!")
                processed = True

            except Exception as e:
                messages.error(request, f"Failed to process Excel file: {e}")
                processed = True

        # Manual form mapping (unchanged)
        co_id = request.POST.get('co')
        po_id = request.POST.get('po')
        level = request.POST.get('level')
        if co_id and po_id and level:
            try:
                co = CO.objects.get(id=co_id, course_id=course_id)
                po = PO.objects.get(id=po_id)
                COPOMapping.objects.create(
                    co=co,
                    po=po,
                    level=level
                )
                messages.success(request, "CO-PO mapping added successfully!")
                processed = True
            except (CO.DoesNotExist, PO.DoesNotExist):
                messages.error(request, "Invalid CO or PO selected.")
                processed = True

        if processed:
            return redirect('add_mapping')
        else:
            messages.warning(request, "No mapping data provided.")

    cos = CO.objects.filter(course__user=request.user)
    pos = PO.objects.all()
    courses = Course.objects.filter(user=request.user)
    return render(request, 'mapping.html', {'cos': cos, 'pos': pos, 'courses': courses})


def upload_marks(request):
    courses = Course.objects.all()

    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        course_id = request.POST.get('course_id')

        try:
            course = Course.objects.get(id=course_id)
            cos = {co.number.strip().upper(): co for co in CO.objects.filter(course=course)}
        except Course.DoesNotExist:
            messages.error(request, "Invalid course selected.")
            return redirect('upload_marks')

        try:
            wb = load_workbook(excel_file)
            sheet = wb.active

            # ✅ Skip merged evaluation row (1st row), read header from row 2
            header = [str(cell.value).strip().upper() for cell in sheet[2]]
            co_headers = header[2:]  # Skip Roll No and Name
        except Exception as e:
            messages.error(request, f"Failed to read Excel file: {e}")
            return redirect('upload_marks')

        created_count = 0
        updated_count = 0
        
        # Initialize dictionaries to accumulate CO totals across all students
        co_totals = defaultdict(lambda: {'total_obtained': 0.0, 'total_max': 0.0, 'student_count': 0})
        processed_students = []

        # ✅ Start reading from row 3
        for row in sheet.iter_rows(min_row=3, values_only=True):
            if not row or row[0] is None:
                continue

            roll_no, name, *marks = row
            roll_no = str(roll_no).strip()
            name = str(name).strip()

            if not roll_no or not name:
                continue

            student, _ = Student.objects.get_or_create(
                roll_number=roll_no,
                defaults={'name': name}
            )
            if student.name != name:
                student.name = name
                student.save()
            
            processed_students.append(student)

            # ✅ 1. Sum all marks by CO number (run for every student)
            co_marks_sum = defaultdict(float)
            for i, co_name in enumerate(co_headers):
                co_key = co_name.strip().upper()
                if i < len(marks) and marks[i] is not None:
                    try:
                        co_marks_sum[co_key] += float(marks[i])
                    except (ValueError, TypeError):
                        continue

            # ✅ 2. Save per-CO StudentMark rows and accumulate totals
            total_obtained_for_student = 0.0
            total_max_for_student = 0.0

            for co_key, total_obtained in co_marks_sum.items():
                co = cos.get(co_key)
                if co:
                    try:
                        obtained_val = float(total_obtained)
                    except (ValueError, TypeError):
                        obtained_val = 0.0

                    # Calculate attainment level
                    if co.max_score > 0:
                        percentage = (obtained_val / co.max_score) * 100
                        if percentage > 70:
                            attainment_level = 3
                        elif percentage > 50:
                            attainment_level = 2
                        else:
                            attainment_level = 1
                    else:
                        attainment_level = 1

                    # Save per-CO StudentMark with attainment level
                    sm_obj, sm_created = StudentMark.objects.update_or_create(
                        course=course,
                        co=co,
                        student=student,
                        defaults={
                            'obtained_marks': obtained_val,
                            'total_marks': co.max_score,
                            'attainment_level': attainment_level
                        }
                    )

                    # Accumulate totals for StudentTotal
                    try:
                        total_obtained_for_student += float(obtained_val)
                        total_max_for_student += float(co.max_score)
                    except (ValueError, TypeError):
                        continue

                    # Accumulate totals for CO aggregate scores
                    co_totals[co_key]['total_obtained'] += obtained_val
                    co_totals[co_key]['total_max'] += co.max_score
                    co_totals[co_key]['student_count'] += 1

            # ✅ 3. Update or create StudentTotal for this student and course
            if total_max_for_student > 0:
                st_obj, st_created = StudentTotal.objects.update_or_create(
                    course=course,
                    student=student,
                    defaults={
                        'total_obtained': total_obtained_for_student,
                        'total_max': total_max_for_student
                    }
                )
                if st_created:
                    created_count += 1
                else:
                    updated_count += 1

        # ✅ 4. Update CO aggregate scores across all students
        for co_key, totals in co_totals.items():
            co = cos.get(co_key)
            if co:
                COAggregateScore.objects.update_or_create(
                    course=course,
                    co=co,
                    defaults={
                        'total_obtained_marks': totals['total_obtained'],
                        'total_max_marks': totals['total_max'],
                        'student_count': totals['student_count']
                    }
                )

        # ✅ 5. Calculate mode attainment level for each CO and store in COAttainment
        for co in cos.values():
            student_marks = StudentMark.objects.filter(course=course, co=co)
            attainment_levels = [sm.attainment_level for sm in student_marks if sm.attainment_level is not None]
            
            if attainment_levels:
                try:
                    # Calculate mode (most frequent attainment level)
                    mode_level = mode(attainment_levels)
                except StatisticsError:
                    # If there's no unique mode, use the average
                    mode_level = round(sum(attainment_levels) / len(attainment_levels))
                
                # Store the mode attainment level in COAttainment
                COAttainment.objects.update_or_create(
                    course=course,
                    co=co,
                    defaults={'level_avg': mode_level}
                )

        messages.success(request, f"{created_count} new records, {updated_count} updated. CO attainment levels calculated and stored.")
        return redirect('upload_marks')

    return render(request, 'upload_marks.html', {'courses': courses})

@login_required
def co_attainment_view(request):
    selected_course_id = request.GET.get('course_id')
    courses = Course.objects.filter(user=request.user)
    attainment_data = []
    selected_course = None
    overall_level_avg = 0

    if selected_course_id:
        selected_course = get_object_or_404(Course, id=selected_course_id, user=request.user)
        cos = CO.objects.filter(course=selected_course)

        total_level = 0
        co_count = 0

        for co in cos:
            student_marks = StudentMark.objects.filter(course=selected_course, co=co)
            levels = []

            for sm in student_marks:
                if sm.total_marks > 0:
                    percent = (sm.obtained_marks / sm.total_marks) * 100
                    if percent >= 60:
                        levels.append(3)
                    elif percent >= 40:
                        levels.append(2)
                    else:
                        levels.append(1)

            avg_level = round(sum(levels) / len(levels), 2) if levels else 0

            # ✅ Save CO attainment
            COAttainment.objects.update_or_create(
                course=selected_course,
                co=co,
                defaults={'level_avg': avg_level}
            )

            attainment_data.append({
                'co_number': co.number,
                'co_description': co.description,
                'level': avg_level
            })

            total_level += avg_level
            co_count += 1

        if co_count > 0:
            overall_level_avg = round(total_level / co_count, 2)

    return render(request, 'co_attainment.html', {
        'courses': courses,
        'selected_course': selected_course,
        'attainment_data': attainment_data,
        'average_level': overall_level_avg
    })

def calculate_po_attainment(request):
    selected_course_id = request.GET.get('course_id')
    courses = Course.objects.filter(user=request.user)
    po_scores = {}
    selected_course = None

    if selected_course_id:
        selected_course = get_object_or_404(Course, id=selected_course_id, user=request.user)
        pos = PO.objects.all()
        
        # Get CO attainments for this course
        co_attainments = {att.co.id: att.level_avg for att in COAttainment.objects.filter(course=selected_course)}
        
        # Get all CO-PO mappings for this course's COs
        course_cos = CO.objects.filter(course=selected_course)
        mappings = COPOMapping.objects.filter(co__in=course_cos).select_related("co", "po")

        # Group mappings by PO for efficient processing
        po_mappings = defaultdict(list)
        for mapping in mappings:
            po_mappings[mapping.po_id].append(mapping)

        for po in pos:
            total_weighted_score = 0
            total_weight = 0

            # Get mappings for this PO
            mappings_for_po = po_mappings.get(po.id, [])
            
            for mapping in mappings_for_po:
                co_id = mapping.co_id
                attainment = co_attainments.get(co_id, 0)  # Default to 0 if missing
                
                if attainment > 0:  # Only include if there's attainment data
                    # Weighted score: attainment level * mapping level
                    weighted_score = attainment * mapping.level
                    total_weighted_score += weighted_score
                    total_weight += mapping.level  # Use mapping level as weight

            # Calculate PO attainment score (weighted average)
            if total_weight > 0:
                # Normalize to percentage (0-100) scale
                # Max possible: 3 (attainment) * 3 (mapping) = 9
                # So we divide by 9 and multiply by 100 to get percentage
                po_score = round((total_weighted_score / total_weight) * (100 / 3), 2)
            else:
                po_score = 0
                
            po_scores[po.number] = po_score
            
            # Store PO attainment in database
            POAttainment.objects.update_or_create(
                course=selected_course,
                po=po,
                defaults={'attainment_score': po_score}
            )

    return render(request, 'po_attainment.html', {
        'courses': courses,
        'selected_course': selected_course,
        'po_scores': po_scores
    })
